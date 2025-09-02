import openpyxl
wb = openpyxl.load_workbook('example3.xlsx')
print(type(wb))

#lsut of sheet names in workbook
print(wb.sheetnames)

sheet = wb['Sheet3']
print(sheet)

sheet_active = wb.active
print(sheet_active)

import openpyxl
wb = openpyxl.load_workbook('example3.xlsx')
sheet = wb['Sheet1']
print(sheet.max_row)
print(sheet.max_column)
print(sheet['A1'].value)

c=sheet['B1']
print(c.value)
print(f'Row {c.row}, Column {c.column} is {c.value}')

print(sheet.cell(row=1, column =2).value)

for i in range(1,8,2):
    print(i, sheet.cell(row = i, column=2).value)

